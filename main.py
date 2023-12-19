import os
import sys
import numpy as np
import openpyxl
import json
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QRadioButton, QPushButton, \
    QInputDialog, QColorDialog, QDesktopWidget, QGridLayout, QTableView, QDialog, QMessageBox, QFileDialog
from PyQt5.QtGui import QFont, QPalette, QColor, QPixmap, QPainter, QBrush
from PyQt5.QtCore import Qt, QAbstractTableModel
from PyQt5.QtWidgets import QMainWindow, QPushButton
from sklearn.cluster import KMeans
import pandas as pd
from sklearn.exceptions import ConvergenceWarning

class TemperamentTestApp(QWidget):
    def __init__(self, questions_and_answers, num_questions):
        super().__init__()
        self.questions_and_answers = questions_and_answers
        self.answers = []
        self.current_question = 0
        self.num_questions = num_questions
        self.finished = False
        self.color_button = None
        self.next_button = None
        self.exit_button = None
        self.initUI()
        self.user_id = self.get_user_id()

    def initUI(self):
        self.layout = QVBoxLayout()
        self.layout.setAlignment(Qt.AlignCenter)
        self.setStyleSheet("color: white")

        # Змінити фон
        self.color_button = QPushButton("Змінити фон")
        self.color_button.setFont(QFont("Times New Roman", 12))
        self.color_button.setMaximumSize(200, 30)
        self.color_button.clicked.connect(self.change_background_color)

        # Встановлення позиції кнопки "Фон" зліва зверху
        top_layout = QHBoxLayout()
        top_layout.addWidget(self.color_button)
        top_layout.setAlignment(Qt.AlignCenter)
        self.layout.addLayout(top_layout)

        self.open_results_button = QPushButton("Відкрити результати тесту")
        self.open_results_button.setFont(QFont("Times New Roman", 12))
        self.open_results_button.setMaximumSize(200, 30)
        self.open_results_button.clicked.connect(self.open_test_results)

        self.open_clusters_button = QPushButton("Відкрити результати кластеризації")
        self.open_clusters_button.setFont(QFont("Times New Roman", 12))
        self.open_clusters_button.setMaximumSize(250, 30)
        self.open_clusters_button.clicked.connect(self.open_cluster_results)

        button_layout = QVBoxLayout()  # Вертикальне вирівнювання
        button_layout.addWidget(self.open_results_button, alignment=Qt.AlignHCenter | Qt.AlignTop)
        button_layout.addWidget(self.open_clusters_button, alignment=Qt.AlignHCenter | Qt.AlignTop)
        self.layout.addLayout(button_layout)

        self.question_label = QLabel(self.questions_and_answers[self.current_question]["question"])
        self.question_label.setFont(QFont("Times New Roman", 16))
        self.question_label.setStyleSheet("color: white;")
        self.layout.addWidget(self.question_label)

        self.radio_buttons = []
        for i, answer in enumerate(self.questions_and_answers[self.current_question]["answers"]):
            radio_button = QRadioButton(f"{answer}")
            radio_button.setFont(QFont("Times New Roman", 14))
            radio_button.setStyleSheet("QRadioButton { color: white; }")  # Set text color to white for radio buttons
            self.radio_buttons.append(radio_button)
            self.layout.addWidget(radio_button)

        self.hlayout = QHBoxLayout()
        self.next_button = QPushButton("Наступне питання")
        self.next_button.setFont(QFont("Times New Roman", 14))
        self.next_button.setEnabled(False)

        self.next_button.clicked.connect(self.next_question)
        self.hlayout.addWidget(self.next_button)
        self.layout.addLayout(self.hlayout)

        # Кнопка "Вихід"
        self.exit_button = QPushButton("Вихід")
        self.exit_button.setFont(QFont("Times New Roman", 14))
        self.exit_button.setMaximumSize(200, 30)
        self.exit_button.clicked.connect(self.close_application)

        # Встановлення вирівнювання кнопки "Вихід" по середині горизонтально
        exit_button_layout = QHBoxLayout()
        exit_button_layout.addWidget(self.exit_button)
        exit_button_layout.setAlignment(Qt.AlignHCenter)
        self.layout.addLayout(exit_button_layout)

        self.setLayout(self.layout)

        # Перевірка вибору варіанта перед активацією кнопки
        for radio_button in self.radio_buttons:
            radio_button.toggled.connect(self.check_selection)

        # Встановлення стилів для кнопок та елементів інтерфейсу
        self.color_button.setStyleSheet("QPushButton { background-color: blue; }")
        self.next_button.setStyleSheet("QPushButton { background-color: lightgreen; }")
        self.exit_button.setStyleSheet("QPushButton { background-color: red; }")
        self.question_label.setStyleSheet("QLabel { color: white; }")
        self.open_results_button.setStyleSheet("QPushButton { background-color: blue; }")
        self.open_clusters_button.setStyleSheet("QPushButton { background-color: blue; }")

        self.accuracy_button = QPushButton("Точність методу k means")
        self.accuracy_button.setFont(QFont("Times New Roman", 12))
        self.accuracy_button.setMaximumSize(250, 30)
        self.accuracy_button.clicked.connect(self.compare_results)

        # Set blue background color for the accuracy button
        self.accuracy_button.setStyleSheet("QPushButton { background-color: blue; }")
        button_layout.addWidget(self.accuracy_button, alignment=Qt.AlignHCenter | Qt.AlignTop)
    def close_application(self):
        sys.exit()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.drawPixmap(self.rect(), QPixmap("C:\\Users\\super\\PycharmProjects\\pythonProject1\\bc.jpg"))

    def change_background_color(self):
        image_path, _ = QFileDialog.getOpenFileName(self, "Select Image", "", "Image Files (*.png *.jpg *.bmp *.gif)")
        if image_path:
            self.set_background_image(image_path)
        else:
            # Set a default background image
            default_image_path = "C:\\Users\\super\\PycharmProjects\\pythonProject1\\default_background.jpg"
            self.set_background_image(default_image_path)

    def set_background_image(self, image_path=None):
        if image_path is None:
            image_path = "C:\\Users\\super\\PycharmProjects\\pythonProject1\\default_background.jpg"

        pixmap = QPixmap(image_path)
        self.setAutoFillBackground(True)
        pal = self.palette()
        pal.setBrush(QPalette.Window, QBrush(pixmap))
        self.setPalette(pal)
    def check_selection(self):
        # Перевірка, чи вибрано хоча б один варіант
        selected = any(radio_button.isChecked() for radio_button in self.radio_buttons)
        self.next_button.setEnabled(selected)

    def open_test_results(self):
        self.open_table("temperament_test_results.xlsx")

    def compare_results(self):
        try:
            # Зчитайте дані з файлів
            test_results = pd.read_excel("temperament_test_results.xlsx")
            cluster_results = pd.read_excel("cluster_results.xlsx")

            # Порівняйте рядки та знайдіть відсоток співпадань
            match_percentage = self.compare_dataframes(test_results.iloc[1:], cluster_results.iloc[1:])

            # Виведіть результат
            result_text = f"Точність методу k means: {match_percentage:.2f}%"
            QMessageBox.information(self, "Результат порівняння", result_text)
        except Exception as e:
            QMessageBox.critical(self, "Помилка", f"Помилка при порівнянні результатів: {str(e)}")

    def compare_dataframes(self, df1, df2):
        try:
            # Заповніть пропущені значення нулями
            df1 = df1.fillna(0)
            df2 = df2.fillna(0)

            # Порівняння рядків з 2 по 5 і обчислення відсотка співпадань
            matches = (df1.values[1:60] - 1) == df2.values[1:60]
            match_percentage = (matches.sum().sum() / matches.size) * 100
            return match_percentage
        except Exception as e:
            raise e

    def open_cluster_results(self):
        self.open_table("cluster_results.xlsx")

    def open_table(self, filename):
        try:
            # Завантаження даних з Excel-файлу в DataFrame
            data = pd.read_excel(filename)

            # Створення вікна для відображення даних
            table_window = QDialog(self)
            table_window.setWindowTitle("Таблиця результатів")
            table_window.setGeometry(100, 100, 1280, 720)

            # Відображення даних у вікні
            table_view = QTableView(table_window)
            table_view.setModel(self.pandasModel(data))  # Виправлено тут
            table_view.setGeometry(10, 10, 1280, 720)

            # Показ вікна
            table_window.exec_()
        except Exception as e:
            QMessageBox.critical(self, "Помилка", f"Помилка при відкритті таблиці: {str(e)}")

    # Клас для відображення даних pandas у QTableView
    class pandasModel(QAbstractTableModel):
        def __init__(self, data):
            QAbstractTableModel.__init__(self)
            self.data = data

        def rowCount(self, parent=None):
            return len(self.data.values)

        def columnCount(self, parent=None):
            return self.data.columns.size

        def data(self, index, role=Qt.DisplayRole):
            if index.isValid() and role == Qt.DisplayRole:
                # Замість використання індексів за замовчуванням, використовуйте назви стовпців і індексів
                value = self.data.iloc[index.row(), index.column()]

                # Перевірте, чи значення - ціле число, і відобразіть його як ціле число
                if isinstance(value, (int, np.int64, np.float64)) and value.is_integer():
                    return str(int(value))

                # Перевірте, чи значення - NaN, і замініть його на порожній рядок
                if pd.isna(value):
                    return ""

                # В іншому випадку, відобразіть значення як текст
                return str(value)

            return None

        def headerData(self, section, orientation, role=Qt.DisplayRole):
            if role == Qt.DisplayRole:
                if orientation == Qt.Horizontal:
                    # Відображення "ID" для першого стовпця, а іншіх - "Питання 1", "Питання 2" і т.д.
                    if section == 0:
                        return "ID"
                    else:
                        return f"Питання {section}"
                elif orientation == Qt.Vertical:
                    # Відображення "Відповіді 1", "Відповіді 2" і т.д. для вертикальної осі
                    return f"Відповіді {section + 1}"
            return None

    def next_question(self):
        selected_answer = None
        for i, radio_button in enumerate(self.radio_buttons):
            if radio_button.isChecked():
                selected_answer = i
                break

        if selected_answer is not None:
            self.answers.append(selected_answer)

        self.current_question += 1

        if self.current_question < self.num_questions:
            self.question_label.setText(self.questions_and_answers[self.current_question]["question"])
            for i, radio_button in enumerate(self.radio_buttons):
                radio_button.setText(f"{self.questions_and_answers[self.current_question]['answers'][i]}")
                radio_button.setChecked(False)
        else:
            self.calculate_and_display_result()

    def get_user_id(self):
        try:
            # Зчитайте номер користувача з файлу
            with open("user_id.json", "r") as file:
                user_id_data = json.load(file)
                return user_id_data["user_id"]
        except FileNotFoundError:
            # Якщо файл не знайдено, почнемо з номера 1
            return 1

    def save_user_id(self):
        # Збережіть збільшений номер користувача в файл
        user_id_data = {"user_id": self.user_id + 1}
        with open("user_id.json", "w") as file:
            json.dump(user_id_data, file)

    def perform_clustering(self, data, num_clusters):
        kmeans = KMeans(n_clusters=num_clusters, random_state=42, n_init=10)  # Додано параметр n_init
        clusters = kmeans.fit_predict(data)
        return clusters

    def calculate_and_display_result(self):
        self.finished = True
        temperament_percentages = calculate_temperament_score(self.answers)
        result_text = "Результати діагностики темпераменту:\n"
        result_text += f"Флегматик: {temperament_percentages[0]:.2f}%\n"
        result_text += f"Сангвіник: {temperament_percentages[1]:.2f}%\n"
        result_text += f"Меланхолік: {temperament_percentages[2]:.2f}%\n"
        result_text += f"Холерик: {temperament_percentages[3]:.2f}%\n"

        # Визначення темпераменту на основі результатів
        temperament_types = ["Флегматик", "Сангвіник", "Меланхолік", "Холерик"]
        max_percentage = max(temperament_percentages)
        temperament_type = temperament_types[temperament_percentages.index(max_percentage)]

        # Додавання пояснення до основного темпераменту
        explanation = "По вашим відповідям, вам тип темперамента скоріш за все це - " + temperament_type + ".\n"
        if temperament_type == "Флегматик":
            explanation += "Темперамент у класифікації Гіппократа. Людину флегматичного темпераменту можна охарактеризувати як вайлувату, зі стійкими прагненнями і більш-менш постійним настроєм, і слабким зовнішнім виразом душевних станів.\n"
            explanation += "Невдале виховання може сприяти формуванню у флегматика таких негативних рис, як млявість, збідненість і слабкість емоцій, схильність до виконання лише звичних дій.\n"
        elif temperament_type == "Сангвіник":
            explanation += "Темперамент сангвініка характеризується високою активністю, життєрадісністю, бажанням спілкуватися з іншими та схильністю до оптимізму. Сангвінік може швидко відновлювати свої сили та відмінно переносити стресові ситуації.\n"
        elif temperament_type == "Меланхолік":
            explanation += "Меланхолічний темперамент характеризується високим рівнем чутливості, роздумливістю та схильністю до міркувань. Меланхолік може ставити високі вимоги до себе та інших, іноді схильний до депресії та тривоги.\n"
        elif temperament_type == "Холерик":
            explanation += "Холеричний темперамент характеризується високим рівнем активності, рішучістю та схильністю до лідерства. Холерик може швидко брати на себе відповідальність і бути наполегливим у досягненні своїх цілей.\n"

        result_text += explanation

        # Розрахунок і відображення підтипу темпераменту
        subtemperament = self.calculate_subtemperament(temperament_percentages)
        result_text += f"Підтип темпераменту: {subtemperament}\n"

        # Кластеризація даних
        data_for_clustering = pd.DataFrame({"answer": self.answers})
        clusters = self.perform_clustering(data_for_clustering, num_clusters=4)

        if clusters is not None:
            # Збереження результатів кластеризації в новому ексель-файлі
            self.save_clusters_to_excel(clusters)
        else:
            print("Clustering failed.")


        # Додавання пояснення до підтипу темпераменту
        subtemperament_explanation = ""
        if subtemperament == "Екстроверт":
            subtemperament_explanation += "Екстраверт - це людина, яка насолоджується спілкуванням з іншими та схильна до активності в групових ситуаціях. Вона отримує енергію від спілкування з оточуючими."
        elif subtemperament == "Інтроверт":
            subtemperament_explanation += "Інтроверт - це людина, яка більше віддає перевагу самотності або спілкуванню в невеликому колі. Вона отримує енергію від самотності та внутрішнього світу."
        elif subtemperament == "Невротичник":
            subtemperament_explanation += "Невротичник - це людина, яка схильна до тривожності, стресу та емоційних коливань. Вона може реагувати на стресові ситуації занадто сильно."
        elif subtemperament == "Стабільний":
            subtemperament_explanation += "Стабільний - це людина, яка здатна добре контролювати свої емоції та справляється зі стресом. Вона зазвичай залишається спокійною в різних ситуаціях."
        elif subtemperament == "Невідомий":
            subtemperament_explanation += "Ваш підтип темпераменту є невідомим. Ви унікальна людина, і ваш темперамент може не вписуватися у звичайні шаблони класифікації. Це означає, що ви маєте унікальну комбінацію рис і особистісних якостей, яка важко піддається класифікації."

        result_text += f"Пояснення підтипу темпераменту:\n{subtemperament_explanation}"

        for radio_button in self.radio_buttons:
            radio_button.hide()
        self.next_button.hide()
        self.question_label.setText(result_text)
        self.save_results_to_excel()
        self.save_user_id()

    def save_clusters_to_excel(self, clusters):
        try:
            file_exists = os.path.isfile("cluster_results.xlsx")

            # Відкрийте або створіть ексель-файл для результатів кластеризації
            if file_exists:
                workbook = openpyxl.load_workbook("cluster_results.xlsx")
                sheet = workbook.active
            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Cluster Results"
                sheet.append(["ID"] + [f"Question {i}" for i in range(1, len(self.answers) + 1)])

            # Запишіть дані про кластери
            sheet.append([self.user_id] + clusters.tolist())

            # Збереження ексель-файлу
            workbook.save("cluster_results.xlsx")
        except Exception as e:
            print(f"An error occurred while saving cluster results: {e}")

    def save_results_to_excel(self):
        try:
            # Спробуйте відкрити існуючий ексель-файл
            workbook = openpyxl.load_workbook("temperament_test_results.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            # Якщо файл не знайдено, створіть новий
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Results"

        # Запис номера користувача
        sheet.cell(row=self.user_id + 1, column=1, value=self.user_id)

        # Запис даних в ексель
        for i, answer in enumerate(self.answers):
            sheet.cell(row=self.user_id + 1, column=i + 2, value=answer + 0)

        # Збереження ексель-файлу
        workbook.save("temperament_test_results.xlsx")
    def calculate_subtemperament(self, temperament_percentages):
        # Знайдіть два найвищих відсотки темпераменту
        top_two_percentages = sorted(temperament_percentages, reverse=True)[:2]

        subtemperament = "Невідомий"

        if top_two_percentages[0] == temperament_percentages[0]:
            if top_two_percentages[1] == temperament_percentages[2]:
                # Флегматик та меланхолік
                subtemperament = "Інтроверт"
            elif top_two_percentages[1] == temperament_percentages[3]:
                # Холерик та меланхолік
                subtemperament = "Невротик"
        elif top_two_percentages[0] == temperament_percentages[1]:
            if top_two_percentages[1] == temperament_percentages[2]:
                # Холерик та сангвіник
                subtemperament = "Екстраверт"
        elif top_two_percentages[0] == temperament_percentages[2] and top_two_percentages[1] == temperament_percentages[
            3]:
            # Сангвіник та флегматик
            subtemperament = "Стабільний"

        return subtemperament


    def reset(self, num_questions):
        self.current_question = 0
        self.answers = []
        self.finished = False
        for radio_button in self.radio_buttons:
            radio_button.show()
            radio_button.setChecked(False)
        self.next_button.show()
        self.question_label.setText(self.questions_and_answers[self.current_question]["question"])
        self.num_questions = num_questions


def calculate_temperament_score(answers):
    temperament_scores = [0, 0, 0, 0]
    for answer in answers:
        temperament_scores[answer] += 1

    total_questions = len(answers)
    temperament_percentages = [score / total_questions * 100 for score in temperament_scores]
    return temperament_percentages


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyleSheet("color: white;")

    num_questions = 10
    questions_and_answers = [
        {
            "question": "Питання №1\nНаскільки часто ви відчуваєте себе пригніченим або сердитим без явних причин?",
            "answers": ["Як правило, я відчуваю стабільність та спокій.",
                        "Я зазвичай налаштований на позитив та енергію, і відчуваю радість від взаємодії з оточуючим світом.",
                        "Я схильний до внутрішніх переживань та можу відчувати себе пригніченим чи сердитим без видимої причини.",
                        "Я часто маю високий рівень енергії та можу реагувати сильно на події, навіть якщо немає конкретної причини"]
        },
        {
            "question": "Питання №2\nЯк ви оцінюєте свою реакцію на несподівані емоційні випробування?",
            "answers": ["Зазвичай я спокійний і легко вирішую такі ситуації.",
                        "Я відношусь до цього оптимістично і намагаюсь підтримувати позитивний настрій.",
                        "Ці емоційні випробування можуть впливати на моє самопочуття, і я звик вивчати їх глибше.",
                        "Я відразу берусь за дію, щоб вирішити проблему і відновити рівновагу."]
        },
        {
            "question": "Питання №3\nЯк ви зазвичай реагуєте на стресові ситуації?",
            "answers": ["Я стараюся залишатися спокійним і знаходити розумні рішення.",
                        "Моя перша реакція - шукати позитив та впорядковувати ситуацію з усмішкою",
                        "Я можу переживати глибокі емоції, але намагаюсь їх розуміти та вирішувати.",
                        "Я відразу берусь за дію, щоб вирішити проблему та відчути контроль"]
        },
        {
            "question": "Питання №4\nЯк ви описали б свою здатність виявляти емпатію?",
            "answers": ["Я можу сприймати емоції інших, іноді це впливає на моє власне настрій.",
                        "Мені легко співпереживати та вислуховувати інших, ставлюся до цього веселим настроєм.",
                        "Я глибоко відчуваю емоції інших і можу відповідати співчуттям та розумінням.",
                        "Моя емпатія зазвичай виявляється в конкретних діях, щоб допомогти вирішити проблему."]
        },
        {
            "question": "Питання №5\nЯк ви реагуєте на невизначеність та невизначені завдання?",
            "answers": ["Я зазвичай залишаюся спокійним, шукаючи стабільність та дієві рішення.",
                        "Це для мене можливість знайти нові можливості та пригоди.",
                        "Невизначеність може викликати стрес, але я стараюся структурувати і планувати для зменшення невизначеності.",
                        "Я приймаю це як виклик і намагаюсь швидко знайти ефективні рішення."]
        },
        {
            "question": "Питання №6\nЯк ви ставитеся до критики або відмови?",
            "answers": ["Я сприймаю це з спокоєм і шукаю можливості для вдосконалення.",
                        "Це може торкнутися мого настрою, але я швидко відновлюю оптимізм та продовжую діяти.",
                        "Це може викликати внутрішні переживання, але я стараюся взяти це як шанс вдосконалитися.",
                        "Я можу відреагувати визначено, але потім відразу переходжу до дії, щоб покращити ситуацію."]
        },
        {
            "question": "Питання №7\nЯк ви ставитеся до невдач і помилок?",
            "answers": ["Я стараюся зрозуміти причини невдачі та вивчити з них уроки для майбутнього.",
                        "Я розглядаю це як частину великої пригоди, де невдача - це лише новий початок.",
                        "Невдачі можуть впливати на мої емоції, але я стараюся аналізувати їх і шукати можливості для вдосконалення.",
                        "Я відразу рухаюсь вперед, шукаючи нові можливості та стратегії після невдачі."]
        },
        {
            "question": "Питання №8\nЯк ви реагуєте на зміни в житті, особливо якщо вони неочікувані?",
            "answers": ["Я зазвичай стаю спокійним і шукаю оптимальні шляхи адаптації до нових обставин.",
                        "Для мене зміни - це можливість для нових пригод і позитивних змін.",
                        "Я можу відчувати себе напруженим від змін, але намагаюсь розглядати їх як можливість для особистого зростання.",
                        "Я відразу берусь за дію, щоб ефективно адаптуватися до нових обставин."]
        },
        {
            "question": "Питання №9\nЯк ви вирішуєте конфлікти в особистих відносинах?",
            "answers": ["Я впорядковую свої емоції і намагаюсь знайти компромісні рішення.",
                        "Я ставлються до конфліктів оптимістично і намагаюсь розважити атмосферу, щоб знайти рішення.",
                        "Я відверто висловлюю свої почуття і співпрацюю для пошуку глибинних рішень.",
                        "Я вперто захищаю свої позиції і долаю конфлікти через активну комунікацію та рішучість."]
        },
        {
            "question": "Питання №10\nЯк ви оцінюєте ризики перед прийняттям важливих рішень?",
            "answers": [
                "Я аналізую ризики та намагаюся приймати рішення, які забезпечують стабільність та невеликий ризик.",
                "Я вважаю ризики можливістю для виклику і впорядковую їх так, щоб зробити рішення захопливим.",
                "Я стараюся докладно вивчити всі аспекти та можливі ризики перед тим, як приймати рішення.",
                "Я швидко аналізую ризики та вирішую, як їх мінімізувати для досягнення мети."]
        },
        ]
    ex = TemperamentTestApp(questions_and_answers, num_questions)
    ex.setGeometry(500, 300, 2000, 900)
    ex.setWindowTitle("Діагностика темпераменту")

    pal = QPalette()
    pal.setColor(QPalette.Window, QColor(155, 189, 214))
    pal.setColor(QPalette.WindowText, Qt.black)
    ex.setPalette(pal)

    ex.show()

    ex.user_id = ex.get_user_id()
    ex.save_user_id()

    screen_resolution = app.desktop().screenGeometry()
    screen_width, screen_height = screen_resolution.width(), screen_resolution.height()

    sys.exit(app.exec_())